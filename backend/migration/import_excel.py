"""Migración del Excel baseline de MARCADOS hacia la base de datos.

PLANTILLA — no se ha ejecutado contra el archivo real todavía porque ese
archivo no está disponible en este entorno de desarrollo. Antes de correr
esto contra datos reales:

  1. Confirmar que se está usando el baseline correcto y vigente (el Excel
     es fuente histórica; no se reordena ni se "mejora" antes de migrar).
  2. Ejecutar con --dry-run primero y revisar el reporte.
  3. Nunca fusionar automáticamente registros ambiguos (ver caso Anthony en
     decisiones-diseño.md de la skill marcados-sistema) — este script no
     fusiona nada; si detecta candidatos duplicados por nombre, los reporta
     para decisión humana y sigue de largo.
  4. Los "13 registros sin rastro" documentados en el histórico no se
     inventan ni se reconstruyen aquí — si no están en el Excel, no entran.

Uso previsto:
    python -m migration.import_excel --excel /ruta/al/baseline.xlsx --dry-run
    python -m migration.import_excel --excel /ruta/al/baseline.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import Base, SessionLocal, engine  # noqa: E402
from app.models import Persona  # noqa: E402
from app.services.identidad import siguiente_id_unico  # noqa: E402

# Mapeo columna Excel -> campo Persona. Ajustar a los encabezados reales del
# baseline vigente antes de ejecutar — estos nombres son los documentados en
# la especificación del proyecto, pueden no coincidir carácter por carácter
# con el archivo real.
MAPEO_COLUMNAS = {
    "ID único": "id_unico_origen",
    "Nombres": "nombres",
    "Apellidos": "apellidos",
    "Fecha de nacimiento": "fecha_nacimiento",
    "Edad manual (respaldo)": "edad_manual",
    "Teléfono": "telefono",
    "Género": "genero",
    "Estado": "estado",
    "Encargado/líder": "encargado_lider",
    "Bautizado": "bautizado",
    "Fecha de bautismo": "fecha_bautismo",
    "Estudio bíblico": "estudio_biblico",
    "Instagram": "instagram",
    "Facebook": "facebook",
    "Dirección": "direccion",
    "Contacto de emergencia": "contacto_emergencia",
    "Teléfono de emergencia": "telefono_emergencia",
    "Grupo sanguíneo": "grupo_sanguineo",
    "EPS": "eps",
    "Talla": "talla",
    "Cómo llegó": "como_llego",
    "Fuente de datos": "fuente_datos",
    "Notas": "notas",
    "Servidor": "servidor",
    "Fecha de ingreso al grupo": "fecha_ingreso",
    "Fecha inicio en servicio": "fecha_inicio_servicio",
}


def leer_filas(ruta_excel: Path, hoja: str = "Jóvenes") -> list[dict]:
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb[hoja] if hoja in wb.sheetnames else wb.active
    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    columnas_presentes = [h for h in encabezados if h in MAPEO_COLUMNAS]
    faltantes = set(MAPEO_COLUMNAS) - set(columnas_presentes)
    if faltantes:
        print(f"AVISO: columnas esperadas y no encontradas en el Excel: {sorted(faltantes)}")

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        registro = {}
        for encabezado, valor in zip(encabezados, row):
            campo = MAPEO_COLUMNAS.get(encabezado)
            if campo:
                registro[campo] = valor
        if registro.get("nombres") or registro.get("apellidos"):
            filas.append(registro)
    return filas


def detectar_posibles_duplicados(filas: list[dict]) -> list[tuple[dict, dict]]:
    """Solo detecta y reporta — nunca decide. Comparación simple por nombre
    normalizado; el veredicto de fusión es siempre humano."""
    from app.matching import normalizar

    sospechosos = []
    vistos: dict[str, dict] = {}
    for fila in filas:
        clave = normalizar(f"{fila.get('nombres', '')} {fila.get('apellidos', '')}")
        if clave in vistos:
            sospechosos.append((vistos[clave], fila))
        else:
            vistos[clave] = fila
    return sospechosos


def migrar(ruta_excel: Path, dry_run: bool) -> None:
    filas = leer_filas(ruta_excel)
    print(f"Filas leídas del Excel: {len(filas)}")

    duplicados = detectar_posibles_duplicados(filas)
    if duplicados:
        print(f"\nPosibles duplicados por nombre ({len(duplicados)}) — requieren decisión humana, NO se fusionan:")
        for a, b in duplicados:
            print(f"  - {a.get('nombres')} {a.get('apellidos')}  <->  {b.get('nombres')} {b.get('apellidos')}")

    if dry_run:
        print("\n--dry-run: no se escribió nada en la base de datos.")
        return

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        creados = 0
        for fila in filas:
            fila.pop("id_unico_origen", None)  # se reasigna con el generador propio de la app
            persona = Persona(id_unico=siguiente_id_unico(db), **fila)
            db.add(persona)
            db.flush()
            creados += 1
        db.commit()
        print(f"\nMigradas {creados} personas.")
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, type=Path, help="Ruta al Excel baseline")
    parser.add_argument("--dry-run", action="store_true", help="Solo reportar, no escribir en la BD")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"No existe el archivo: {args.excel}")
        raise SystemExit(1)

    migrar(args.excel, dry_run=args.dry_run)
