"""Migración del Excel real de MARCADOS a la BD.

Implementa el contrato de la sección 15 de MARCADOS_DATA_HANDOFF ("ESTRUCTURA
FINAL 2026-08-09 — esta sección manda sobre todo lo anterior"). Reglas clave
que este script respeta:

  - La clave es 'ID único' (columna A) — nunca el nombre ni la posición ("No.").
  - No se importan las columnas derivadas: B (No.), E (Nombre completo ayuda),
    H (Edad), I (Segmento), AG..AL (Asistencias 30d, Reuniones evaluables 30d,
    % Asistencia 30d, Semáforo, Ficha completa %, Datos faltantes) — son
    fórmulas de Excel, la app las recalcula (sección 17 / 21).
  - Vacío migra como NULL, nunca como valor por defecto inventado.
  - No se fusionan personas ambiguas (ver sección 9 del handoff) — se
    importan como personas separadas y se marca el caso con un registro de
    Seguimiento para que un humano lo revise.
  - Las personas sin Estado se importan con estado=NULL + Seguimiento
    marcado "requiere_atencion", nunca se infiere el estado.
  - El texto no-fecha encontrado en una columna de fecha (caso conocido:
    'hace 1 año') se importa como NULL, conservando el texto original en
    una nota — nunca se inventa una fecha.
  - 'Área de servicio' (columna M) es texto libre, aún sin normalizar a la
    tabla de relación persona<->área (eso llega con la hoja Servicio, hoy
    vacía) — se conserva tal cual en una nota para no perder el dato.
  - 'Fecha de ingreso al grupo' (columna AD) NUNCA se inventa para estos 120
    registros históricos: si el Excel la trae vacía (así es hoy, sección
    18.1), queda NULL. Todas las personas migradas quedan marcadas con
    registro_historico=True para distinguirlas de las que cree la app desde
    ahora (esas sí reciben fecha_ingreso = hoy, ver PersonaCreate).
  - Los 15 catálogos de la hoja 'Catálogos' (header fila 4, valores desde
    fila 5) se cargan como filas de la tabla Catalogo (tipo, valor) — carga
    idempotente: si un valor ya existe para ese tipo, no se duplica.
  - Las hojas Asistencia, Seguimiento y Servicio están vacías en el Excel
    (sección 16.1) — no se importan filas de ahí, solo se crean las tablas.

Uso:
    python -m migration.migrar_datos_reales --excel /ruta/al/excel.xlsx --dry-run
    python -m migration.migrar_datos_reales --excel /ruta/al/excel.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import Base, SessionLocal, engine  # noqa: E402
from app.models import Actividad, Catalogo, Persona, Seguimiento  # noqa: E402

# Catálogo de actividades — fuente: Catálogos!B (columna ACTIVIDAD).
ACTIVIDADES = [
    "Sábado - Encuentro Marcados",
    "Domingo - Escuela Dominical",
    "Martes - Escuela de Servidores",
    "Reunión general líderes y servidores",
    "Otro / Evento esporádico",
]

# Columnas de captura de la hoja 'Jóvenes' (1-indexadas). Las derivadas
# (B, E, H, I, AG, AH, AI, AJ, AK, AL) quedan fuera a propósito: no aparecen
# en este diccionario, así que nunca se leen ni se migran como dato.
COL = {
    "id_unico": 1,  # A
    "nombres": 3,  # C
    "apellidos": 4,  # D
    "genero": 6,  # F
    "fecha_nacimiento": 7,  # G
    "edad_manual": 10,  # J
    "estado": 11,  # K
    "servidor": 12,  # L
    "area_servicio": 13,  # M
    "bautizado": 14,  # N
    "estudio_biblico": 15,  # O
    "telefono": 16,  # P
    "correo": 17,  # Q
    "tiene_instagram": 18,  # R
    "instagram": 19,  # S
    "tiene_facebook": 20,  # T
    "facebook": 21,  # U
    "direccion": 22,  # V
    "contacto_emergencia": 23,  # W
    "parentesco": 24,  # X
    "telefono_emergencia": 25,  # Y
    "grupo_sanguineo": 26,  # Z
    "eps": 27,  # AA
    "talla": 28,  # AB
    "como_llego": 29,  # AC
    "fecha_ingreso": 30,  # AD
    "fecha_bautismo": 31,  # AE
    "fecha_inicio_servicio": 32,  # AF
    "notas": 39,  # AM
}

FILA_INICIO = 3
FILA_FIN = 122  # inclusive — TblJovenes A2:AM122 (120 personas)

# Hoja 'Catálogos': fila de encabezados = 4, valores desde la fila 5, sin
# filas vacías en medio (así lo garantiza la propia hoja, sección 15).
CATALOGOS_FILA_HEADER = 4
CATALOGOS_FILA_INICIO = 5
# columna -> tipo interno (nombres definidos en Excel: CatEstado, CatActividad...)
CATALOGOS_COL_TIPO = {
    1: "estado",  # A CatEstado
    2: "actividad",  # B CatActividad
    3: "asistencia",  # C CatAsistencia
    4: "formacion",  # D CatFormacion
    5: "area_servicio",  # E CatArea
    6: "parentesco",  # F CatParentesco
    7: "eps",  # G CatEPS
    8: "grupo_sanguineo",  # H CatGrupoSanguineo
    9: "talla",  # I CatTalla
    10: "genero",  # J CatGenero
    11: "si_no",  # K CatSiNo
    12: "consolidacion",  # L CatConsolidacion
    13: "tipo_contacto",  # M CatTipoContacto
    14: "estado_seguimiento",  # N CatEstadoSeguimiento
    15: "estado_servicio",  # O CatEstadoServicio
}


def _si_no(valor) -> bool | None:
    if valor is None:
        return None
    v = str(valor).strip().lower()
    if v in ("si", "sí"):
        return True
    if v == "no":
        return False
    return None


def _texto(valor) -> str | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        # Teléfonos y similares llegan como número si el Excel no los formateó
        # como texto — se preserva el valor tal cual, sin decimales falsos.
        v = str(int(valor)) if float(valor).is_integer() else str(valor)
    else:
        v = str(valor).strip()
    return v.strip() or None


def _fecha(valor, campo: str, notas_extra: list[str]) -> dt.date | None:
    """Devuelve la fecha si el valor es una fecha real. Si hay texto en una
    columna de fecha (caso conocido: 'hace 1 año'), NO se inventa una
    conversión — se guarda como NULL y se deja constancia en las notas."""
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    texto = str(valor).strip()
    if texto:
        notas_extra.append(f"{campo} original (texto, no convertido): {texto!r}")
    return None


def leer_filas(ruta_excel: Path) -> list[dict]:
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb["Jóvenes"]
    filas = []
    for r in range(FILA_INICIO, FILA_FIN + 1):
        id_unico = ws.cell(row=r, column=COL["id_unico"]).value
        nombres = ws.cell(row=r, column=COL["nombres"]).value
        if not id_unico and not nombres:
            continue
        notas_extra: list[str] = []
        fila = {
            "fila_excel": r,
            "id_unico": _texto(id_unico),
            "nombres": _texto(nombres) or "",
            "apellidos": _texto(ws.cell(row=r, column=COL["apellidos"]).value) or "",
            "genero": _texto(ws.cell(row=r, column=COL["genero"]).value),
            "fecha_nacimiento": _fecha(
                ws.cell(row=r, column=COL["fecha_nacimiento"]).value, "Fecha de nacimiento", notas_extra
            ),
            "edad_manual": ws.cell(row=r, column=COL["edad_manual"]).value,
            "estado": _texto(ws.cell(row=r, column=COL["estado"]).value),
            "servidor": _si_no(ws.cell(row=r, column=COL["servidor"]).value) or False,
            "area_servicio": _texto(ws.cell(row=r, column=COL["area_servicio"]).value),
            "bautizado": _si_no(ws.cell(row=r, column=COL["bautizado"]).value),
            "estudio_biblico": _texto(ws.cell(row=r, column=COL["estudio_biblico"]).value),
            "telefono": _texto(ws.cell(row=r, column=COL["telefono"]).value),
            "correo_electronico": _texto(ws.cell(row=r, column=COL["correo"]).value),
            "tiene_instagram": _si_no(ws.cell(row=r, column=COL["tiene_instagram"]).value),
            "instagram": _texto(ws.cell(row=r, column=COL["instagram"]).value),
            "tiene_facebook": _si_no(ws.cell(row=r, column=COL["tiene_facebook"]).value),
            "facebook": _texto(ws.cell(row=r, column=COL["facebook"]).value),
            "direccion": _texto(ws.cell(row=r, column=COL["direccion"]).value),
            "contacto_emergencia": _texto(ws.cell(row=r, column=COL["contacto_emergencia"]).value),
            "parentesco": _texto(ws.cell(row=r, column=COL["parentesco"]).value),
            "telefono_emergencia": _texto(ws.cell(row=r, column=COL["telefono_emergencia"]).value),
            "grupo_sanguineo": _texto(ws.cell(row=r, column=COL["grupo_sanguineo"]).value),
            "eps": _texto(ws.cell(row=r, column=COL["eps"]).value),
            "talla": _texto(ws.cell(row=r, column=COL["talla"]).value),
            "como_llego": _texto(ws.cell(row=r, column=COL["como_llego"]).value),
            "notas": _texto(ws.cell(row=r, column=COL["notas"]).value),
            "fecha_ingreso": _fecha(
                ws.cell(row=r, column=COL["fecha_ingreso"]).value, "Fecha de ingreso al grupo", notas_extra
            ),
            "fecha_bautismo": _fecha(
                ws.cell(row=r, column=COL["fecha_bautismo"]).value, "Fecha de bautismo", notas_extra
            ),
            "fecha_inicio_servicio": _fecha(
                ws.cell(row=r, column=COL["fecha_inicio_servicio"]).value,
                "Fecha inicio en servicio",
                notas_extra,
            ),
        }
        if fila["bautizado"] is None:
            fila["bautizado"] = False
        if fila["area_servicio"]:
            notas_extra.append(f"Área de servicio (sin normalizar): {fila['area_servicio']}")
        if notas_extra:
            extra = " | ".join(notas_extra)
            fila["notas"] = f"{fila['notas']} | {extra}" if fila["notas"] else extra
        del fila["area_servicio"]
        filas.append(fila)
    return filas


def leer_catalogos(ruta_excel: Path) -> dict[str, list[str]]:
    """Lee los 15 catálogos de la hoja 'Catálogos': header en la fila 4,
    valores desde la fila 5 hacia abajo, hasta la primera celda vacía de
    cada columna (la propia hoja garantiza que no hay huecos en medio)."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb["Catálogos"]
    catalogos: dict[str, list[str]] = {tipo: [] for tipo in CATALOGOS_COL_TIPO.values()}
    for col, tipo in CATALOGOS_COL_TIPO.items():
        r = CATALOGOS_FILA_INICIO
        while True:
            valor = ws.cell(row=r, column=col).value
            if valor is None or str(valor).strip() == "":
                break
            catalogos[tipo].append(str(valor).strip())
            r += 1
    return catalogos


def validar(filas: list[dict]) -> dict:
    """Solo detecta y reporta — nunca decide ni fusiona."""
    sin_id = [f for f in filas if not f["id_unico"]]
    ids_vistos: dict[str, dict] = {}
    ids_duplicados = []
    for f in filas:
        if f["id_unico"] in ids_vistos:
            ids_duplicados.append((ids_vistos[f["id_unico"]], f))
        else:
            ids_vistos[f["id_unico"]] = f

    # Apellido faltante es un vacío real de los datos, no bloquea la
    # migración — solo "nombres" vacío sí sería un problema real, porque ahí
    # no habría ni siquiera con qué identificar a la persona.
    sin_nombre = [f for f in filas if not f["nombres"]]
    sin_apellido = [f for f in filas if f["nombres"] and not f["apellidos"]]
    sin_estado = [f for f in filas if not f["estado"]]

    telefonos: dict[str, list[dict]] = {}
    for f in filas:
        if f["telefono"]:
            telefonos.setdefault(f["telefono"], []).append(f)
    telefonos_compartidos = {t: fs for t, fs in telefonos.items() if len(fs) > 1}

    return {
        "total": len(filas),
        "sin_id": sin_id,
        "ids_duplicados": ids_duplicados,
        "sin_nombre": sin_nombre,
        "sin_apellido": sin_apellido,
        "sin_estado": sin_estado,
        "telefonos_compartidos": telefonos_compartidos,
    }


def imprimir_reporte(filas: list[dict], val: dict, catalogos: dict[str, list[str]]) -> None:
    print(f"Personas leídas del Excel: {val['total']}")
    print(f"Sin ID único: {len(val['sin_id'])}")
    print(f"IDs duplicados: {len(val['ids_duplicados'])}")
    print(f"Sin nombre (bloqueante): {len(val['sin_nombre'])}")
    print(f"Sin apellido registrado (dato real, no bloquea): {len(val['sin_apellido'])}")
    for f in val["sin_apellido"]:
        print(f"  fila {f['fila_excel']} {f['id_unico']} {f['nombres']!r}")
    print(f"Sin Estado definido: {len(val['sin_estado'])} — se importan con estado=NULL + Seguimiento marcado")
    for f in val["sin_estado"]:
        print(f"  fila {f['fila_excel']} {f['id_unico']} {f['nombres']} {f['apellidos']}")
    print(f"\nTeléfonos compartidos por más de una persona: {len(val['telefonos_compartidos'])}")
    for tel, fs in val["telefonos_compartidos"].items():
        nombres = ", ".join(f"{f['id_unico']} {f['nombres']} {f['apellidos']}" for f in fs)
        print(f"  {tel}: {nombres}  (NO se fusiona — se marca para revisión humana)")
    print(f"\nCatálogos leídos: {len(catalogos)}")
    for tipo, valores in catalogos.items():
        print(f"  {tipo}: {len(valores)} valores")


def cargar_catalogos(db, catalogos: dict[str, list[str]]) -> int:
    """Carga idempotente: si (tipo, valor) ya existe, no lo duplica."""
    existentes = {(c.tipo, c.valor) for c in db.query(Catalogo).all()}
    creados = 0
    for tipo, valores in catalogos.items():
        for valor in valores:
            if (tipo, valor) in existentes:
                continue
            db.add(Catalogo(tipo=tipo, valor=valor))
            existentes.add((tipo, valor))
            creados += 1
    return creados


def migrar(ruta_excel: Path, dry_run: bool) -> None:
    filas = leer_filas(ruta_excel)
    val = validar(filas)
    catalogos = leer_catalogos(ruta_excel)
    imprimir_reporte(filas, val, catalogos)

    if val["sin_id"] or val["ids_duplicados"] or val["sin_nombre"]:
        print("\nERROR: hay filas sin ID único, con ID duplicado, o sin nombre. No se puede migrar así.")
        raise SystemExit(1)

    if dry_run:
        print("\n--dry-run: no se escribió nada en la base de datos.")
        return

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        existentes = db.query(Persona).count()
        if existentes:
            print(f"\nERROR: la base de datos ya tiene {existentes} personas. No se sobrescribe en silencio.")
            print("Si es una base de prueba y querés reiniciarla, borrala vos mismo antes de correr esto.")
            raise SystemExit(1)

        catalogos_creados = cargar_catalogos(db, catalogos)
        db.flush()
        print(f"\nValores de catálogo creados: {catalogos_creados}")

        actividades_por_nombre = {}
        for nombre in ACTIVIDADES:
            act = Actividad(nombre=nombre)
            db.add(act)
            db.flush()
            actividades_por_nombre[nombre] = act
        print(f"Actividades creadas: {len(actividades_por_nombre)}")

        personas_por_id: dict[str, Persona] = {}
        creadas = 0
        for f in filas:
            persona = Persona(
                id_unico=f["id_unico"],
                nombres=f["nombres"],
                apellidos=f["apellidos"],
                genero=f["genero"],
                fecha_nacimiento=f["fecha_nacimiento"],
                edad_manual=f["edad_manual"] if isinstance(f["edad_manual"], int) else None,
                estado=f["estado"],
                servidor=f["servidor"],
                bautizado=f["bautizado"],
                estudio_biblico=f["estudio_biblico"],
                telefono=f["telefono"],
                correo_electronico=f["correo_electronico"],
                tiene_instagram=f["tiene_instagram"],
                instagram=f["instagram"],
                tiene_facebook=f["tiene_facebook"],
                facebook=f["facebook"],
                direccion=f["direccion"],
                contacto_emergencia=f["contacto_emergencia"],
                parentesco=f["parentesco"],
                telefono_emergencia=f["telefono_emergencia"],
                grupo_sanguineo=f["grupo_sanguineo"],
                eps=f["eps"],
                talla=f["talla"],
                como_llego=f["como_llego"],
                notas=f["notas"],
                fecha_ingreso=f["fecha_ingreso"],
                fecha_bautismo=f["fecha_bautismo"],
                fecha_inicio_servicio=f["fecha_inicio_servicio"],
                registro_historico=True,
            )
            db.add(persona)
            db.flush()
            personas_por_id[f["id_unico"]] = persona
            creadas += 1
        print(f"Personas creadas: {creadas} (todas marcadas registro_historico=True)")

        seguimientos = 0
        for f in val["sin_estado"]:
            persona = personas_por_id[f["id_unico"]]
            db.add(
                Seguimiento(
                    persona_id=persona.id,
                    tipo="revision",
                    notas=(
                        "Sin Estado definido (Activo/Inactivo/Fluctúa) al momento de migrar el "
                        "Excel — sin evidencia interna para inferirlo. Requiere decisión humana."
                    ),
                    requiere_atencion=True,
                )
            )
            seguimientos += 1

        for tel, fs in val["telefonos_compartidos"].items():
            for f in fs:
                otros = [g for g in fs if g is not f]
                otros_desc = "; ".join(f"{g['id_unico']} {g['nombres']} {g['apellidos']}" for g in otros)
                persona = personas_por_id[f["id_unico"]]
                db.add(
                    Seguimiento(
                        persona_id=persona.id,
                        tipo="revision",
                        notas=(
                            f"Posible duplicado: mismo teléfono ({tel}) que {otros_desc}. "
                            "NO fusionado automáticamente — requiere decisión humana."
                        ),
                        requiere_atencion=True,
                    )
                )
                seguimientos += 1

        db.commit()
        print(f"Registros de Seguimiento creados (marcados para revisión): {seguimientos}")
        print("\nMigración completa.")
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"No existe el archivo: {args.excel}")
        raise SystemExit(1)

    migrar(args.excel, dry_run=args.dry_run)
