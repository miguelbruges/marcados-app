from datetime import date

import openpyxl
import pytest
from openpyxl.worksheet.table import Table

from app.models import Actividad, Asistencia, Evento, Persona
from app.services.exportador_excel import generar_export


def _construir_plantilla_sintetica(ruta):
    """Plantilla mínima, 100% sintética (sin ningún dato real), con la
    misma forma estructural que el libro real: dos hojas con tablas de
    Excel, una fórmula en una columna 'derivada' para probar que se copia
    tal cual a las filas nuevas, y un rango inicial de una sola fila de
    datos para forzar la extensión de la tabla en el test."""
    wb = openpyxl.Workbook()
    ws_jov = wb.active
    ws_jov.title = "Jóvenes"
    for col in range(1, 40):
        ws_jov.cell(row=2, column=col).value = f"Col{col}"
    ws_jov.cell(row=3, column=37).value = "=1+1"  # columna derivada de ejemplo (como AK)
    ws_jov.add_table(Table(displayName="TblJovenes", ref="A2:AM3"))

    ws_asis = wb.create_sheet("Asistencia")
    for col in range(1, 9):
        ws_asis.cell(row=3, column=col).value = f"Col{col}"
    ws_asis.cell(row=4, column=8).value = "=IF(1=1,1,0)"  # columna derivada (como H)
    ws_asis.add_table(Table(displayName="TblAsistencia", ref="A3:H4"))

    wb.save(ruta)


@pytest.fixture()
def plantilla(tmp_path):
    ruta = tmp_path / "plantilla_sintetica.xlsx"
    _construir_plantilla_sintetica(ruta)
    return ruta


def test_export_llena_jovenes_y_extiende_la_tabla_copiando_formulas(db_session, plantilla):
    for i in range(3):
        db_session.add(
            Persona(
                id_unico=f"MAR-00000{i}",
                nombres=f"Nombre{i}",
                apellidos=f"Apellido{i}",
                genero="Femenino",
                telefono="300",
                bautizado=True,
            )
        )
    db_session.commit()

    buffer = generar_export(db_session, plantilla)
    wb = openpyxl.load_workbook(buffer)
    ws = wb["Jóvenes"]

    assert ws.tables["TblJovenes"].ref == "A2:AM5"  # header(2) + 3 filas de datos
    assert [ws.cell(row=r, column=1).value for r in (3, 4, 5)] == ["MAR-000000", "MAR-000001", "MAR-000002"]
    assert ws.cell(row=3, column=3).value == "Nombre0"
    assert ws.cell(row=3, column=12).value == "No"  # servidor=False -> "No"
    assert ws.cell(row=3, column=14).value == "Sí"  # bautizado=True -> "Sí"

    # la formula de la columna derivada (AK=37) se copió tal cual a las filas nuevas
    assert ws.cell(row=3, column=37).value == "=1+1"
    assert ws.cell(row=4, column=37).value == "=1+1"
    assert ws.cell(row=5, column=37).value == "=1+1"


def test_export_llena_asistencia_y_no_toca_la_columna_formula(db_session, plantilla):
    persona = Persona(id_unico="MAR-000001", nombres="Ana", apellidos="Perez")
    actividad = Actividad(nombre="Sábado - Encuentro Marcados")
    db_session.add_all([persona, actividad])
    db_session.flush()

    evento = Evento(actividad_id=actividad.id, nombre="Sábado 2026-01-10", fecha=date(2026, 1, 10))
    db_session.add(evento)
    db_session.flush()
    db_session.add(Asistencia(persona_id=persona.id, evento_id=evento.id, presente=True))
    db_session.commit()

    buffer = generar_export(db_session, plantilla)
    wb = openpyxl.load_workbook(buffer)
    ws = wb["Asistencia"]

    assert ws.cell(row=4, column=1).value.date() == date(2026, 1, 10)  # openpyxl relee fechas como datetime
    assert ws.cell(row=4, column=2).value == "Sábado - Encuentro Marcados"
    assert ws.cell(row=4, column=4).value == "Ana Perez"
    assert ws.cell(row=4, column=5).value == "Asistió"
    # la columna 8 (ID persona resuelto) es fórmula: no se sobrescribe con dato
    assert ws.cell(row=4, column=8).value == "=IF(1=1,1,0)"


def test_export_endpoint_sin_plantilla_configurada_da_503(client, auth_headers):
    from app import config

    original = config.settings.excel_template_path
    config.settings.excel_template_path = None
    try:
        resp = client.get("/export/excel", headers=auth_headers)
        assert resp.status_code == 503
    finally:
        config.settings.excel_template_path = original


def test_estado_plantilla_sin_ninguna_configurada(client, auth_headers):
    resp = client.get("/export/plantilla/estado", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.json()["configurada"] is False


def test_subir_plantilla_y_luego_descargar_usa_la_guardada(client, auth_headers, plantilla):
    with open(plantilla, "rb") as f:
        resp = client.post(
            "/export/plantilla",
            files={"archivo": ("plantilla.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_headers,
        )
    assert resp.status_code == 204

    estado = client.get("/export/plantilla/estado", headers=auth_headers).json()
    assert estado["configurada"] is True
    assert estado["nombre_archivo"] == "plantilla.xlsx"

    resp = client.get("/export/excel", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_subir_plantilla_reemplaza_la_anterior(client, auth_headers, plantilla):
    with open(plantilla, "rb") as f:
        client.post("/export/plantilla", files={"archivo": ("primera.xlsx", f, "application/octet-stream")}, headers=auth_headers)
    with open(plantilla, "rb") as f:
        client.post("/export/plantilla", files={"archivo": ("segunda.xlsx", f, "application/octet-stream")}, headers=auth_headers)

    estado = client.get("/export/plantilla/estado", headers=auth_headers).json()
    assert estado["nombre_archivo"] == "segunda.xlsx"


def test_subir_plantilla_archivo_invalido_da_400(client, auth_headers):
    resp = client.post(
        "/export/plantilla",
        files={"archivo": ("no-es-excel.xlsx", b"esto no es un excel", "application/octet-stream")},
        headers=auth_headers,
    )
    assert resp.status_code == 400


def test_subir_plantilla_no_admin_da_403(client, db_session, plantilla):
    from app.models import RolUsuario, Usuario
    from app.security import hash_password

    lider = Usuario(nombre="Lider", email="lider9@marcadosapp.dev", password_hash=hash_password("x"), rol=RolUsuario.LIDER)
    db_session.add(lider)
    db_session.commit()
    token = client.post("/auth/login", json={"email": "lider9@marcadosapp.dev", "password": "x"}).json()["access_token"]

    with open(plantilla, "rb") as f:
        resp = client.post(
            "/export/plantilla",
            files={"archivo": ("plantilla.xlsx", f, "application/octet-stream")},
            headers={"Authorization": f"Bearer {token}"},
        )
    assert resp.status_code == 403


def test_export_endpoint_requiere_admin(client, db_session):
    from app.models import RolUsuario, Usuario
    from app.security import hash_password

    lider = Usuario(nombre="Lider", email="lider2@marcadosapp.dev", password_hash=hash_password("x"), rol=RolUsuario.LIDER)
    db_session.add(lider)
    db_session.commit()
    token = client.post("/auth/login", json={"email": "lider2@marcadosapp.dev", "password": "x"}).json()["access_token"]

    resp = client.get("/export/excel", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 403


# --- Recálculo al abrir (pedido del usuario, 2026-08-24) ---
# openpyxl no evalúa fórmulas y descarta los valores ya calculados de la
# plantilla al reguardar: medido sobre el export real, de 2333 fórmulas
# quedaban 0 con valor. Sin la marca de recálculo el archivo se abre en
# blanco en cualquier lector que no recalcule solo.

def _calc_pr(ruta_o_buffer):
    import re
    import zipfile

    z = zipfile.ZipFile(ruta_o_buffer)
    xml = z.read("xl/workbook.xml").decode()
    m = re.search(r"<calcPr[^>]*/?>", xml)
    return m.group(0) if m else None


def test_export_pide_recalculo_al_abrir(plantilla, db_session):
    salida = generar_export(db_session, plantilla)
    assert 'fullCalcOnLoad="1"' in _calc_pr(salida)


def test_export_fuerza_recalculo_aunque_la_plantilla_lo_tenga_apagado(tmp_path, db_session):
    """El peor caso: una plantilla guardada con fullCalcOnLoad="0" daría un
    archivo vacío Y sin recalcular. La bandera se fija en el código, no se
    hereda de la plantilla."""
    from openpyxl.workbook.properties import CalcProperties

    ruta = tmp_path / "plantilla_sin_recalculo.xlsx"
    _construir_plantilla_sintetica(ruta)
    wb = openpyxl.load_workbook(ruta)
    wb.calculation = CalcProperties(calcId=191029, fullCalcOnLoad=False)
    wb.save(ruta)
    assert 'fullCalcOnLoad="0"' in _calc_pr(ruta)

    salida = generar_export(db_session, ruta)
    assert 'fullCalcOnLoad="1"' in _calc_pr(salida)


def test_export_fuerza_recalculo_si_la_plantilla_no_tiene_calcPr(tmp_path, db_session):
    ruta = tmp_path / "plantilla_sin_calcpr.xlsx"
    _construir_plantilla_sintetica(ruta)
    wb = openpyxl.load_workbook(ruta)
    wb.calculation = None
    wb.save(ruta)

    salida = generar_export(db_session, ruta)
    assert 'fullCalcOnLoad="1"' in _calc_pr(salida)
