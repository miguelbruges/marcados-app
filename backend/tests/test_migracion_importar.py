import io

import openpyxl


def _construir_excel(filas, con_catalogos=True):
    """filas: list of dicts con las claves que use cada test — solo llena
    las columnas mencionadas, no hace falta reconstruir la fila completa.
    Mismo mapeo de columnas real (A=id_unico, C=nombres, D=apellidos,
    F=genero, K=estado, L=servidor, N=bautizado, P=telefono)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jóvenes"
    for i, fila in enumerate(filas):
        r = 3 + i
        ws.cell(row=r, column=1).value = fila.get("id_unico")
        ws.cell(row=r, column=3).value = fila.get("nombres", "")
        ws.cell(row=r, column=4).value = fila.get("apellidos", "")
        if "genero" in fila:
            ws.cell(row=r, column=6).value = fila["genero"]
        if "estado" in fila:
            ws.cell(row=r, column=11).value = fila["estado"]
        if "servidor" in fila:
            ws.cell(row=r, column=12).value = "Sí" if fila["servidor"] else "No"
        if "bautizado" in fila:
            ws.cell(row=r, column=14).value = "Sí" if fila["bautizado"] else "No"
        if "telefono" in fila:
            ws.cell(row=r, column=16).value = fila["telefono"]

    if con_catalogos:
        ws_cat = wb.create_sheet("Catálogos")
        ws_cat.cell(row=4, column=1).value = "ESTADO"
        ws_cat.cell(row=5, column=1).value = "Activo"
        ws_cat.cell(row=6, column=1).value = "Inactivo"
        ws_cat.cell(row=4, column=10).value = "GÉNERO"
        ws_cat.cell(row=5, column=10).value = "Masculino"
        ws_cat.cell(row=6, column=10).value = "Femenino"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _importar(client, headers, filas, confirmar=False):
    archivo = _construir_excel(filas)
    return client.post(
        f"/migracion/importar?confirmar={'true' if confirmar else 'false'}",
        files={"archivo": ("import.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


# --- Crear (base vacía o IDs que no existen todavía) ---

def test_vista_previa_no_escribe_nada(client, auth_headers, db_session):
    from app.models import Persona

    resp = _importar(
        client, auth_headers,
        [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "estado": "Activo"},
         {"id_unico": "MAR-000002", "nombres": "Camila", "apellidos": "Rodriguez", "estado": "Activo"}],
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is True
    assert len(body["a_crear"]) == 2
    assert db_session.query(Persona).count() == 0


def test_confirmar_crea_personas_nuevas(client, auth_headers, db_session):
    from app.models import Persona

    resp = _importar(
        client, auth_headers,
        [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "estado": "Activo"},
         {"id_unico": "MAR-000002", "nombres": "Camila", "apellidos": "Rodriguez", "estado": "Activo"}],
        confirmar=True,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is False
    assert body["resultado"]["personas_creadas"] == 2
    assert db_session.query(Persona).count() == 2
    creadas = db_session.query(Persona).all()
    assert all(p.registro_historico for p in creadas)


def test_sin_estado_al_crear_genera_seguimiento_de_revision(client, auth_headers, db_session):
    from app.models import Persona, Seguimiento

    resp = _importar(
        client, auth_headers,
        [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez"}],
        confirmar=True,
    )
    assert resp.status_code == 200
    persona = db_session.query(Persona).filter_by(id_unico="MAR-000001").one()
    seg = db_session.query(Seguimiento).filter_by(persona_id=persona.id).first()
    assert seg is not None
    assert seg.requiere_atencion is True


def test_telefono_compartido_entre_nuevas_genera_seguimiento(client, auth_headers, db_session):
    from app.models import Persona, Seguimiento

    resp = _importar(
        client, auth_headers,
        [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "estado": "Activo", "telefono": "3000000000"},
         {"id_unico": "MAR-000002", "nombres": "Camila", "apellidos": "Rodriguez", "estado": "Activo", "telefono": "3000000000"}],
        confirmar=True,
    )
    assert resp.status_code == 200
    p1 = db_session.query(Persona).filter_by(id_unico="MAR-000001").one()
    p2 = db_session.query(Persona).filter_by(id_unico="MAR-000002").one()
    assert db_session.query(Seguimiento).filter_by(persona_id=p1.id).count() == 1
    assert db_session.query(Seguimiento).filter_by(persona_id=p2.id).count() == 1


# --- Actualizar (IDs que ya existen) ---

def test_preview_detecta_cambios_sin_escribir(client, auth_headers, db_session):
    from app.models import Persona

    p = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="3000000000")
    db_session.add(p)
    db_session.commit()

    resp = _importar(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "3009999999"}])
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is True
    assert len(body["a_actualizar"]) == 1
    assert body["a_actualizar"][0]["id_unico"] == "MAR-000001"
    assert body["a_actualizar"][0]["campos_cambiados"]["telefono"] == {"antes": "3000000000", "despues": "3009999999"}
    assert body["a_crear"] == []

    db_session.refresh(p)
    assert p.telefono == "3000000000"  # nada escrito todavía


def test_confirmar_aplica_los_cambios_y_registra_bitacora(client, auth_headers, db_session):
    from app.models import Bitacora, Persona

    p = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="3000000000")
    db_session.add(p)
    db_session.commit()

    resp = _importar(
        client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "3009999999"}], confirmar=True
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["resultado"]["personas_actualizadas"] == 1
    assert body["resultado"]["personas_creadas"] == 0

    db_session.refresh(p)
    assert p.telefono == "3009999999"

    bitacora = db_session.query(Bitacora).filter(Bitacora.registro_id == p.id, Bitacora.campo == "telefono").first()
    assert bitacora is not None
    assert bitacora.valor_anterior == "3000000000"
    assert bitacora.valor_nuevo == "3009999999"


def test_no_toca_a_quien_no_aparece_en_el_excel(client, auth_headers, db_session):
    from app.models import Persona

    p1 = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="111")
    p2 = Persona(id_unico="MAR-000002", nombres="Camila", apellidos="Rodriguez", telefono="222")
    db_session.add_all([p1, p2])
    db_session.commit()

    resp = _importar(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "999"}], confirmar=True)
    assert resp.status_code == 200

    db_session.refresh(p2)
    assert p2.telefono == "222"


def test_no_vacia_el_nombre_por_una_celda_en_blanco(client, auth_headers, db_session):
    from app.models import Persona

    p = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez")
    db_session.add(p)
    db_session.commit()

    resp = _importar(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "", "apellidos": "", "telefono": "999"}], confirmar=True)
    assert resp.status_code == 200

    db_session.refresh(p)
    assert p.nombres == "Sofia"
    assert p.apellidos == "Hernandez"
    assert p.telefono == "999"


def test_celda_en_blanco_no_borra_datos_existentes(client, auth_headers, db_session):
    from datetime import date

    from app.models import Persona

    p = Persona(
        id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez",
        telefono="3000000000", fecha_ingreso=date(2026, 1, 1), direccion="Calle 1",
    )
    db_session.add(p)
    db_session.commit()

    # Excel armado a mano sin llenar telefono/fecha_ingreso/direccion (a
    # propósito, simulando un archivo incompleto) — nada de eso debe
    # borrarse, solo lo que sí venga con un valor real.
    resp = _importar(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "estado": "Activo"}], confirmar=True)
    assert resp.status_code == 200

    db_session.refresh(p)
    assert p.telefono == "3000000000"
    assert p.fecha_ingreso == date(2026, 1, 1)
    assert p.direccion == "Calle 1"
    assert p.estado == "Activo"  # esto sí vino con valor real, se aplica


def test_personas_sin_mencionar_en_el_excel_se_reportan(client, auth_headers, db_session):
    from app.models import Persona

    db_session.add_all(
        [
            Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez"),
            Persona(id_unico="MAR-000002", nombres="Camila", apellidos="Rodriguez"),
        ]
    )
    db_session.commit()

    resp = _importar(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez"}])
    body = resp.json()
    assert [p["id_unico"] for p in body["personas_sin_mencionar"]] == ["MAR-000002"]


# --- Crear y actualizar mezclado en un mismo archivo ---

def test_mismo_archivo_crea_y_actualiza_a_la_vez(client, auth_headers, db_session):
    from app.models import Persona

    existente = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="111")
    db_session.add(existente)
    db_session.commit()

    resp = _importar(
        client, auth_headers,
        [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "999"},
         {"id_unico": "MAR-000002", "nombres": "Nueva", "apellidos": "Persona", "estado": "Activo"}],
        confirmar=True,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["resultado"]["personas_creadas"] == 1
    assert body["resultado"]["personas_actualizadas"] == 1
    assert db_session.query(Persona).count() == 2

    db_session.refresh(existente)
    assert existente.telefono == "999"


# --- Validación / bloqueo ---

def test_id_duplicado_bloquea_incluso_en_vista_previa(client, auth_headers):
    resp = _importar(
        client, auth_headers,
        [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez"},
         {"id_unico": "MAR-000001", "nombres": "Camila", "apellidos": "Rodriguez"}],
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "error" in body
    assert len(body["ids_duplicados"]) == 1


def test_sin_nombre_bloquea(client, auth_headers):
    body = _importar(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "", "apellidos": "SinNombre"}]).json()
    assert "error" in body


def test_archivo_no_xlsx_rechazado(client, auth_headers):
    resp = client.post(
        "/migracion/importar",
        files={"archivo": ("prueba.txt", io.BytesIO(b"no es un excel"), "text/plain")},
        headers=auth_headers,
    )
    assert resp.status_code == 422


def test_solo_admin_puede_importar(client, db_session):
    from app.models import RolUsuario, Usuario
    from app.security import hash_password

    lider = Usuario(nombre="Lider", email="lider5@marcadosapp.dev", password_hash=hash_password("x"), rol=RolUsuario.LIDER)
    db_session.add(lider)
    db_session.commit()
    token = client.post("/auth/login", json={"email": "lider5@marcadosapp.dev", "password": "x"}).json()["access_token"]

    resp = _importar(client, {"Authorization": f"Bearer {token}"}, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez"}])
    assert resp.status_code == 403
