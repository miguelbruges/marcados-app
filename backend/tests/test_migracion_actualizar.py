import io

import openpyxl


def _construir_excel_actualizacion(filas):
    """filas: list of dicts. Solo llena las columnas que los tests de este
    archivo necesitan — no hace falta reconstruir la fila completa."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jóvenes"
    for i, fila in enumerate(filas):
        r = 3 + i
        ws.cell(row=r, column=1).value = fila.get("id_unico")
        ws.cell(row=r, column=3).value = fila.get("nombres", "")
        ws.cell(row=r, column=4).value = fila.get("apellidos", "")
        if "telefono" in fila:
            ws.cell(row=r, column=16).value = fila["telefono"]
        if "estado" in fila:
            ws.cell(row=r, column=11).value = fila["estado"]
        if "bautizado" in fila:
            ws.cell(row=r, column=14).value = "Sí" if fila["bautizado"] else "No"
        if "servidor" in fila:
            ws.cell(row=r, column=12).value = "Sí" if fila["servidor"] else "No"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _subir(client, headers, filas, confirmar=False):
    archivo = _construir_excel_actualizacion(filas)
    return client.post(
        f"/migracion/excel-actualizar?confirmar={'true' if confirmar else 'false'}",
        files={"archivo": ("actualizacion.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


def test_preview_detecta_cambios_sin_escribir(client, auth_headers, db_session):
    from app.models import Persona

    p = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="3000000000")
    db_session.add(p)
    db_session.commit()

    resp = _subir(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "3009999999"}])
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is True
    assert len(body["personas_con_cambios"]) == 1
    assert body["personas_con_cambios"][0]["id_unico"] == "MAR-000001"
    assert body["personas_con_cambios"][0]["campos_cambiados"]["telefono"] == {"antes": "3000000000", "despues": "3009999999"}

    db_session.refresh(p)
    assert p.telefono == "3000000000"  # nada escrito todavía


def test_confirmar_aplica_los_cambios_y_registra_bitacora(client, auth_headers, db_session):
    from app.models import Bitacora, Persona

    p = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="3000000000")
    db_session.add(p)
    db_session.commit()

    resp = _subir(
        client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "3009999999"}], confirmar=True
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is False
    assert body["resultado"]["personas_actualizadas"] == 1

    db_session.refresh(p)
    assert p.telefono == "3009999999"

    bitacora = db_session.query(Bitacora).filter(Bitacora.registro_id == p.id, Bitacora.campo == "telefono").first()
    assert bitacora is not None
    assert bitacora.valor_anterior == "3000000000"
    assert bitacora.valor_nuevo == "3009999999"


def test_no_toca_a_quien_no_aparece_en_el_excel(client, auth_headers, db_session):
    from app.models import Persona

    p1 = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez", telefono="111")
    p2 = Persona(id_unico="MAR-000002", nombres="Camila", apellidos="Rodriguez", telefono="222")
    db_session.add_all([p1, p2])
    db_session.commit()

    resp = _subir(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "telefono": "999"}], confirmar=True)
    assert resp.status_code == 200

    db_session.refresh(p2)
    assert p2.telefono == "222"


def test_reporta_id_unico_no_encontrado_sin_fallar(client, auth_headers, db_session):
    resp = _subir(client, auth_headers, [{"id_unico": "MAR-999999", "nombres": "Fantasma", "apellidos": "Nadie"}])
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["filas_sin_coincidencia"]) == 1
    assert body["filas_sin_coincidencia"][0]["id_unico"] == "MAR-999999"


def test_no_vacia_el_nombre_por_una_celda_en_blanco(client, auth_headers, db_session):
    from app.models import Persona

    p = Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez")
    db_session.add(p)
    db_session.commit()

    resp = _subir(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "", "apellidos": "", "telefono": "999"}], confirmar=True)
    assert resp.status_code == 200

    db_session.refresh(p)
    assert p.nombres == "Sofia"
    assert p.apellidos == "Hernandez"
    assert p.telefono == "999"


def test_celda_en_blanco_no_borra_datos_existentes(client, auth_headers, db_session):
    from datetime import date

    from app.models import Persona

    p = Persona(
        id_unico="MAR-000001",
        nombres="Sofia",
        apellidos="Hernandez",
        telefono="3000000000",
        fecha_ingreso=date(2026, 1, 1),
        direccion="Calle 1",
    )
    db_session.add(p)
    db_session.commit()

    # Excel armado a mano sin llenar telefono/fecha_ingreso/direccion (a
    # propósito, simulando un archivo incompleto) — nada de eso debe
    # borrarse, solo lo que sí venga con un valor real.
    resp = _subir(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez", "estado": "Activo"}], confirmar=True)
    assert resp.status_code == 200

    db_session.refresh(p)
    assert p.telefono == "3000000000"
    assert p.fecha_ingreso == date(2026, 1, 1)
    assert p.direccion == "Calle 1"
    assert p.estado == "Activo"  # esto sí vino con valor real, se aplica


def test_personas_sin_mencionar_en_el_excel_se_reportan(client, auth_headers, db_session):
    from app.models import Persona

    db_session.add_all(
        [
            Persona(id_unico="MAR-000001", nombres="Sofia", apellidos="Hernandez"),
            Persona(id_unico="MAR-000002", nombres="Camila", apellidos="Rodriguez"),
        ]
    )
    db_session.commit()

    resp = _subir(client, auth_headers, [{"id_unico": "MAR-000001", "nombres": "Sofia", "apellidos": "Hernandez"}])
    body = resp.json()
    assert [p["id_unico"] for p in body["personas_sin_mencionar"]] == ["MAR-000002"]


def test_actualizar_desde_excel_requiere_admin(client, db_session):
    from app.models import RolUsuario, Usuario
    from app.security import hash_password

    lider = Usuario(nombre="Lider", email="lider10@marcadosapp.dev", password_hash=hash_password("x"), rol=RolUsuario.LIDER)
    db_session.add(lider)
    db_session.commit()
    token = client.post("/auth/login", json={"email": "lider10@marcadosapp.dev", "password": "x"}).json()["access_token"]

    resp = _subir(client, {"Authorization": f"Bearer {token}"}, [{"id_unico": "MAR-000001", "nombres": "X", "apellidos": "Y"}])
    assert resp.status_code == 403
