import io

import openpyxl
import pytest


def _construir_excel_sintetico(*, con_id_repetido=False, sin_nombre=False):
    """Libro sintético mínimo (sin ningún dato real) con la forma de las
    hojas 'Jóvenes' y 'Catálogos' que exige el mapeo de columnas real."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jóvenes"

    filas = [
        # id_unico, nombres, apellidos, genero
        ("MAR-000001", "Sofia", "Hernandez", "Femenino"),
        ("MAR-000002" if not con_id_repetido else "MAR-000001", "Camila", "Rodriguez", "Femenino"),
    ]
    if sin_nombre:
        filas.append(("MAR-000003", "", "SinNombre", "Masculino"))

    for i, (id_unico, nombres, apellidos, genero) in enumerate(filas):
        r = 3 + i
        ws.cell(row=r, column=1).value = id_unico  # A
        ws.cell(row=r, column=3).value = nombres  # C
        ws.cell(row=r, column=4).value = apellidos  # D
        ws.cell(row=r, column=6).value = genero  # F
        ws.cell(row=r, column=11).value = "Activo"  # K estado

    ws_cat = wb.create_sheet("Catálogos")
    ws_cat.cell(row=4, column=1).value = "ESTADO"
    ws_cat.cell(row=5, column=1).value = "Activo"
    ws_cat.cell(row=6, column=1).value = "Inactivo"
    ws_cat.cell(row=4, column=10).value = "GÉNERO"
    ws_cat.cell(row=5, column=10).value = "Masculino"
    ws_cat.cell(row=6, column=10).value = "Femenino"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_vista_previa_no_escribe_nada(client, auth_headers, db_session):
    from app.models import Persona

    archivo = _construir_excel_sintetico()
    resp = client.post(
        "/migracion/excel",
        files={"archivo": ("prueba.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is True
    assert body["total_leidas"] == 2
    assert db_session.query(Persona).count() == 0


def test_confirmar_migra_de_verdad(client, auth_headers, db_session):
    from app.models import Persona

    archivo = _construir_excel_sintetico()
    resp = client.post(
        "/migracion/excel?confirmar=true",
        files={"archivo": ("prueba.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["vista_previa"] is False
    assert body["resultado"]["personas_creadas"] == 2
    assert db_session.query(Persona).count() == 2
    creadas = db_session.query(Persona).all()
    assert all(p.registro_historico for p in creadas)


def test_confirmar_se_niega_si_ya_hay_personas(client, auth_headers):
    archivo1 = _construir_excel_sintetico()
    client.post(
        "/migracion/excel?confirmar=true",
        files={"archivo": ("prueba.xlsx", archivo1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )

    archivo2 = _construir_excel_sintetico()
    resp = client.post(
        "/migracion/excel?confirmar=true",
        files={"archivo": ("prueba2.xlsx", archivo2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert resp.status_code == 409


def test_id_duplicado_bloquea_incluso_en_vista_previa(client, auth_headers):
    archivo = _construir_excel_sintetico(con_id_repetido=True)
    resp = client.post(
        "/migracion/excel",
        files={"archivo": ("prueba.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "error" in body
    assert body["ids_duplicados"] == 1


def test_sin_nombre_bloquea(client, auth_headers):
    archivo = _construir_excel_sintetico(sin_nombre=True)
    resp = client.post(
        "/migracion/excel",
        files={"archivo": ("prueba.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    body = resp.json()
    assert "error" in body


def test_archivo_no_xlsx_rechazado(client, auth_headers):
    resp = client.post(
        "/migracion/excel",
        files={"archivo": ("prueba.txt", io.BytesIO(b"no es un excel"), "text/plain")},
        headers=auth_headers,
    )
    assert resp.status_code == 422


def test_solo_admin_puede_migrar(client, db_session):
    from app.models import RolUsuario, Usuario
    from app.security import hash_password

    lider = Usuario(nombre="Lider", email="lider5@marcadosapp.dev", password_hash=hash_password("x"), rol=RolUsuario.LIDER)
    db_session.add(lider)
    db_session.commit()
    token = client.post("/auth/login", json={"email": "lider5@marcadosapp.dev", "password": "x"}).json()["access_token"]

    archivo = _construir_excel_sintetico()
    resp = client.post(
        "/migracion/excel",
        files={"archivo": ("prueba.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 403
