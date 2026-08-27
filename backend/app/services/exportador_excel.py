"""Export a Excel usando el libro real como plantilla — sección 16.2 del
handoff. No genera el .xlsx desde cero: carga la plantilla real (con su
diseño, catálogos, nombres definidos y TODAS las fórmulas), sobrescribe
solo las filas de datos de las tablas Jóvenes y Asistencia, y ajusta el
rango de cada tabla al número real de filas. Nunca escribe en columnas
derivadas (Edad, Segmento, Ficha completa %, Datos faltantes, ID persona
resuelto...) — Excel las recalcula solo al abrir el archivo.

La plantilla NUNCA vive en este repositorio: contiene datos personales de
menores de edad. Se sube una vez desde Administración y queda guardada en
la base de datos (tabla plantilla_excel) — el disco de Render es efímero,
así que un archivo subido "a mano" al servidor no sobrevive a un redeploy.

Seguimiento y Servicio no se exportan todavía: sus columnas en el Excel
(Responsable, Resultado, Próxima acción / Área de servicio con relación
persona<->área) no tienen un equivalente exacto y ya poblado en el modelo
actual — exportarlas ahora exigiría inventar un mapeo. Se deja pendiente
a propósito en vez de adivinar.
"""

from __future__ import annotations

import io
from copy import copy
from pathlib import Path
from typing import BinaryIO

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Asistencia, Evento, Persona

# Columnas de captura de 'Jóvenes' (sección 15 del handoff, columna -> campo
# del modelo). Mismo mapeo que backend/migration/migrar_datos_reales.py,
# aquí en sentido inverso (de la base de datos hacia el Excel). Las
# derivadas (B,E,H,I,AG..AL) no aparecen: nunca se tocan.
COL_JOVENES: dict[int, str] = {
    1: "id_unico",
    3: "nombres",
    4: "apellidos",
    6: "genero",
    7: "fecha_nacimiento",
    10: "edad_manual",
    11: "estado",
    12: "servidor",
    14: "bautizado",
    15: "estudio_biblico",
    16: "telefono",
    17: "correo_electronico",
    18: "tiene_instagram",
    19: "instagram",
    20: "tiene_facebook",
    21: "facebook",
    22: "direccion",
    23: "contacto_emergencia",
    24: "parentesco",
    25: "telefono_emergencia",
    26: "grupo_sanguineo",
    27: "eps",
    28: "talla",
    29: "como_llego",
    30: "fecha_ingreso",
    31: "fecha_bautismo",
    32: "fecha_inicio_servicio",
    39: "notas",
}

CAMPOS_SI_NO = {"servidor", "bautizado", "tiene_instagram", "tiene_facebook"}


def _valor_excel(persona: Persona, campo: str):
    valor = getattr(persona, campo)
    if campo in CAMPOS_SI_NO:
        if valor is None:
            return None
        return "Sí" if valor else "No"
    return valor


def _asegurar_filas(ws: Worksheet, nombre_tabla: str, filas_necesarias: int) -> None:
    """Extiende (o encoge) la tabla `nombre_tabla` a `filas_necesarias` filas
    de datos, copiando fórmula y estilo de la última fila existente cuando
    hace falta agregar filas nuevas. Las referencias estructuradas de Excel
    (TblX[[#This Row],[Columna]]) no dependen del número de fila, así que
    copiar la fórmula tal cual a la fila nueva es correcto."""
    tabla = ws.tables[nombre_tabla]
    min_col, header_row, max_col, max_row = range_boundaries(tabla.ref)
    filas_actuales = max_row - header_row
    fila_plantilla = max_row

    if filas_necesarias > filas_actuales:
        for i in range(filas_actuales + 1, filas_necesarias + 1):
            nueva_fila = header_row + i
            for col in range(min_col, max_col + 1):
                origen = ws.cell(row=fila_plantilla, column=col)
                destino = ws.cell(row=nueva_fila, column=col)
                if isinstance(origen.value, str) and origen.value.startswith("="):
                    destino.value = origen.value
                if origen.has_style:
                    destino._style = copy(origen._style)

    filas_finales = max(filas_necesarias, 1)  # una tabla de Excel necesita al menos 1 fila de datos
    nueva_max_fila = header_row + filas_finales
    tabla.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{nueva_max_fila}"


def _llenar_jovenes(ws: Worksheet, personas: list[Persona]) -> None:
    _asegurar_filas(ws, "TblJovenes", len(personas))
    fila = 3  # primera fila de datos de TblJovenes (header en la fila 2)
    for persona in personas:
        for col, campo in COL_JOVENES.items():
            ws.cell(row=fila, column=col).value = _valor_excel(persona, campo)
        fila += 1


def _llenar_asistencia(ws: Worksheet, asistencias: list[Asistencia]) -> None:
    _asegurar_filas(ws, "TblAsistencia", len(asistencias))
    fila = 4  # primera fila de datos de TblAsistencia (header en la fila 3)
    for a in asistencias:
        ws.cell(row=fila, column=1).value = a.evento.fecha
        ws.cell(row=fila, column=2).value = a.evento.actividad.nombre
        ws.cell(row=fila, column=3).value = None  # Detalle: solo aplica a 'Otro / esporádico', no lo inventamos
        ws.cell(row=fila, column=4).value = a.persona.nombre_completo
        ws.cell(row=fila, column=5).value = "Asistió" if a.presente else "Inasistió"
        ws.cell(row=fila, column=6).value = None  # Observación: no hay campo equivalente todavía
        ws.cell(row=fila, column=7).value = a.registrado_por.nombre if a.registrado_por else None
        # columna 8 (ID persona resuelto) es fórmula — no se toca
        fila += 1


def generar_export(db: Session, plantilla: Path | BinaryIO) -> io.BytesIO:
    """Devuelve un .xlsx en memoria — nunca se guarda en disco del repo.
    `plantilla` acepta una ruta (uso local) o un archivo en memoria
    (BytesIO con el contenido guardado en la base, uso en producción)."""
    wb = openpyxl.load_workbook(plantilla)  # sin data_only: conserva las fórmulas

    personas = db.scalars(select(Persona).order_by(Persona.id_unico)).all()
    _llenar_jovenes(wb["Jóvenes"], personas)

    asistencias = db.scalars(
        select(Asistencia).join(Evento).order_by(Evento.fecha, Asistencia.persona_id)
    ).all()
    _llenar_asistencia(wb["Asistencia"], asistencias)

    _forzar_recalculo_al_abrir(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _forzar_recalculo_al_abrir(wb: openpyxl.Workbook) -> None:
    """Marca el libro para que Excel recalcule TODO al abrirlo.

    Por qué hace falta (medido sobre el export real que reportó el usuario,
    2026-08-24): openpyxl no evalúa fórmulas, y al reguardar descarta los
    valores ya calculados que traía la plantilla. En números: de las 2333
    fórmulas del libro, la plantilla tenía 1325 con valor guardado y el
    archivo exportado quedaba con 0. Sin esta marca, cualquier programa que
    no recalcule muestra todas esas celdas vacías — los dos paneles, Edad,
    Segmento, % de asistencia, Semáforo, Ficha completa.

    A propósito NO se restauran los valores viejos de la plantilla: fueron
    calculados sobre los datos de ANTES, así que se leerían como cifras
    actuales siendo falsas. En blanco hasta recalcular es honesto; un número
    desactualizado con cara de bueno, no — y acá se toman decisiones sobre
    personas mirando esos números.

    Se fija explícitamente en vez de confiar en lo que traiga la plantilla:
    hoy funciona de casualidad porque openpyxl completa la bandera al
    reescribir calcPr, pero una plantilla con fullCalcOnLoad="0" daría un
    archivo en blanco Y sin recalcular, que es el peor caso posible.
    """
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True
    # calcId=0 es "no sé con qué versión se calculó esto": refuerza el
    # recálculo incluso en lectores que ignoren fullCalcOnLoad.
    wb.calculation.calcId = 0
